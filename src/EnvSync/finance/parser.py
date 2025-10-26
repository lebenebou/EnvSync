
import os
import sys
import re

from tabulate import tabulate
import time
import argparse

from EnvSync import GlobalEnv

REPORTS_DIR = os.path.join(GlobalEnv.ENCRYPTED_PATH, "finance")
CACHED_DIR = os.path.join(REPORTS_DIR, "Cached")
MASTER_EXCEL_FILE = os.path.join(REPORTS_DIR, "master.xlsx")

import datetime
import pandas
from openpyxl import load_workbook
import pdfplumber

from EnvSync.finance.Transaction import Transaction, Series, Currency, TransactionLocation
from EnvSync.finance.helpers import parseDate, parseFloat, tryParseFloat

def cacheSeries(series: Series):

    print(f'Caching series with {len(series.transactions)} transactions...', end=' ', flush=True, file=sys.stderr)

    if len(series.transactions) == 0:
        return

    today: str = datetime.datetime.now().strftime('%Y_%m_%d')
    csvFileName = f'{today}.csv'
    csvFilePath = os.path.join(CACHED_DIR, csvFileName)

    dataFrame = series.toDataFrame()
    print(f'(to {csvFilePath})', flush=True, file=sys.stderr)
    dataFrame.to_csv(csvFilePath, index=False)

def transactionsFromBankAudiPDF(pdfPath: str) -> list[Transaction]:

    print(f'Parsing Bank Audi {os.path.basename(pdfPath)}...', end=' ', flush=True, file=sys.stderr)

    data: list[list[str]] = []
    with pdfplumber.open(pdfPath) as pdf:

        for page in pdf.pages:
            tables = page.extract_tables()

            for table in tables:
                data.extend(table)

    # remove newlines from cells
    for i, row in enumerate(data):
        data[i] = list(map(lambda cell: cell.replace('\n', ' ') if cell else None, row))

    # only keep transaction rows
    firstTransactionIndex = 0
    for i, row in enumerate(data):
        if any(map(lambda cell : cell.lower().count('transaction') if cell else False, row)):
            firstTransactionIndex = i
            break

    data = data[firstTransactionIndex:]

    dataFrame: pandas.DataFrame = pandas.DataFrame(data[1:], columns=data[0])
    
    dataFrame.drop(columns=['Long Description'], inplace=True)

    # Some rows will have half of their values on the next page on the pdf
    for i, row in dataFrame.iterrows():
        if all(row): # skip full row
            continue

        for j in range(len(row)):
            if row.iloc[j] == None:
                continue

            dataFrame.iloc[i-1, j] = dataFrame.iloc[i-1, j] + row.iloc[j] # join with the row above, in the row above
            row.iloc[j] = None # to make the whole row None

    # drop empty
    dataFrame.dropna(how='all', inplace=True)

    transactions: list[Transaction] = []
    for _, row in dataFrame.iterrows():

        t = Transaction()

        t.uniqueId = row['Serial Number'].replace(' ', '')
        t.uniqueId = int(t.uniqueId)

        t.date = parseDate(row['Transaction Date'])
        t.description = str(row['Description'])

        t.credit = parseFloat(row['Credit'])

        if t.credit == 0:
            t.credit = -1 * parseFloat(row['Debit'])

        t.balance = parseFloat(row['Running Balance'])

        t.fillTypeAndLocation()
        t.accountName = 'Bank Audi'
        transactions.append(t)

    print(f'Parsed {len(transactions)} transactions.', flush=True, file=sys.stderr)
    return transactions

def transactionsFromRevolutCSV(csvFilePath: str) -> list[Transaction]:

    print(f'Parsing Revolut {os.path.basename(csvFilePath)}...', end=' ', flush=True, file=sys.stderr)

    dataFrame: pandas.DataFrame = pandas.read_csv(csvFilePath)
    transactions: list[Transaction] = []

    for _, row in dataFrame.iterrows():

        knownStates = ['COMPLETED', 'PENDING', 'REVERTED']
        if row['State'] not in knownStates:
            print(f'[WARN] unknown revolut transaction state: {row["State"]}', flush=True, file=sys.stderr)
            continue

        t = Transaction(currency=row['Currency'])

        t.description = row['Description']

        revolutDateFormat: str = '%Y-%m-%d %H:%M:%S'
        initialDate: str = row['Started Date'] # example: 8/13/2025 11:12

        t.date = parseDate(initialDate, revolutDateFormat)
        time = datetime.datetime.strptime(initialDate, revolutDateFormat)
        t.uniqueId = int(time.timestamp())

        t.credit = parseFloat(row['Amount'])

        if row['State'] == 'REVERTED':
            t.credit = abs(t.credit)

        t.feeAmount = parseFloat(row['Fee'])
        t.balance = parseFloat(row['Balance'])

        t.accountName = 'Revolut'
        t.fillTypeAndLocation()
        t.location = TransactionLocation.france
        t.correctAttributeTypes()

        transactions.append(t)

    print(f'Parsed {len(transactions)} transactions.', flush=True, file=sys.stderr)
    return transactions

def transactionsFromCachedCsv(csvFilePath: str) -> list[Transaction]:

    dataFrame: pandas.DataFrame = pandas.read_csv(csvFilePath)
    transactions: list[Transaction] = []
    
    print(f'Parsing transactions from dataframe {csvFilePath}...', flush=True, file=sys.stderr)

    for _, row in dataFrame.iterrows():
        t = Transaction.fromDataFrameRow(row)
        transactions.append(t)

    return transactions

def getLatestCachedCsvFile(folder: str = REPORTS_DIR) -> str:

    csvFiles = [f for f in os.listdir(CACHED_DIR) if f.endswith('.csv')]
    latestFile = max(csvFiles, key=lambda f: datetime.datetime.strptime(f.strip('.csv'), '%Y_%m_%d'))

    return os.path.join(CACHED_DIR, latestFile)

def printObjectList(objects: list[object], csv: bool = False):

    if len(objects) == 0:
        return

    print(end='\n', flush=True, file=sys.stderr)

    tableContent = [obj.__dict__.values() for obj in objects]
    headers = [key.capitalize() for key in objects[0].__dict__.keys()]

    if csv:
        fullTable: str = tabulate(tableContent, headers=headers, tablefmt='tsv').replace('\t', ',')
        fullTable = re.sub(r'\s*,\s*', r',', fullTable) # remove all spaces in between commas
    else:
        fullTable: str = tabulate(tableContent, headers=headers)

    headerCutOff = 1 if csv else 2
    headerContent = '\n'.join(fullTable.split('\n')[:headerCutOff])
    tableContent = '\n'.join(fullTable.split('\n')[headerCutOff:])

    if not len(tableContent):
        return

    if not csv:
        print(headerContent, file=sys.stderr, end='\n\n') # allow the use of grep while keeping the headers
    else:
        print(headerContent, file=sys.stdout, end='\n')

    time.sleep(0.005) # this is to avoid stderr getting mixed with stdout, force headers to first line

    try:
        print(tableContent, file=sys.stdout)
    except BrokenPipeError: # some commands like "head" will close the pipe early and prevent the program from outputting more lines
        pass

def updateMasterExcelWithNewTransactions(fullData: pandas.DataFrame) -> int:

    print(f'Updating master excel with {len(fullData)} row(s)...', flush=True, file=sys.stderr)

    book = load_workbook(MASTER_EXCEL_FILE)
    sheet = book.active

    # delete all except headers
    sheet.delete_rows(2, sheet.max_row-1)

    for i, row in fullData.iterrows():
        for j, value in enumerate(row, start=1):
            sheet.cell(row=i+2, column=j, value=tryParseFloat(value))

    try:
        book.save(MASTER_EXCEL_FILE)
    except PermissionError:
        print(f'Could not save to file: {MASTER_EXCEL_FILE}', file=sys.stderr)
        print(f'It might be open in excel, please close it', file=sys.stderr)
        return 1

    return 0

if __name__ == '__main__':

    parser = argparse.ArgumentParser(description='Parse an account activity report from Bank Audi')
    parser.add_argument('--csv', action='store_true', default=False, help='Output in csv format')

    limitArg = parser.add_argument_group()
    limitArg.add_argument('--all', action='store_true', help='Show all transactions, overrides --limit', default=False)

    fileType = parser.add_mutually_exclusive_group()
    fileType.add_argument('--refresh', action='store_true', help='Parse and cache "./Reports/audi.pdf"')
    fileType.add_argument('--open', action='store_true', help='Update and open the main transactions excel file')

    filterArg = parser.add_argument_group()
    filterArg.add_argument('--filter', type=str, help='--filter=X: only transactions with X in the description')
    filterArg.add_argument('--type', type=str, help='--type=C: only transactions of type C')
    filterArg.add_argument('--location', type=str, help='--location=L: only transactions in L')

    filterArg.add_argument('--after', type=str, help='--after=dd-mm-yyyy: only transactions after this date', default=None)
    filterArg.add_argument('--before', type=str, help='--before=dd-mm-yyyy: only transactions before this date', default=None)

    parser.add_argument('-c', '--currency', type=str, help='Example: --currency=EUR, convert all transactions to this currency', default='USD')

    args = parser.parse_args()

    transactions: list[Transaction] = []
    series: Series = None

    if args.refresh:
        series = Series('USD', transactionsFromBankAudiPDF(os.path.join(REPORTS_DIR, 'audi.pdf')))
        series.extend(transactionsFromRevolutCSV(os.path.join(REPORTS_DIR, 'revolut.csv')) )
        cacheSeries(series)
    else:
        transactions = transactionsFromCachedCsv(getLatestCachedCsvFile())
        series = Series('USD', transactions)

    if args.open:
        returnCode: int = updateMasterExcelWithNewTransactions(series.toDataFrame())

        if returnCode != 0:
            exit(returnCode)

        os.startfile(MASTER_EXCEL_FILE)
        exit(0)

    if args.filter:
        series.filterBySubstring(args.filter)

    if args.location:
        series.filterByLocation(args.location)

    if args.type:
        series.filterByCategory(args.type)

    dateLowerBound = '01-01-1970'
    dateLowerBound = datetime.datetime.strptime(dateLowerBound, "%d-%m-%Y").date()
    if args.after:
        dateLowerBound = datetime.datetime.strptime(args.after, "%d-%m-%Y").date()

    dateUpperBound = '01-01-2099'
    dateUpperBound = datetime.datetime.strptime(dateUpperBound, "%d-%m-%Y").date()
    if args.before:
        dateUpperBound = datetime.datetime.strptime(args.before, "%d-%m-%Y").date()

    if args.after or args.before:
        series.dateFilter(dateLowerBound, dateUpperBound)

    assert Currency.currencySupported(args.currency), f'Currency not supported: {args.currency}'
    series.convertToCurrency(args.currency)

    if not args.csv:
        series.addTotal()
        series.prepareForPrettyPrint()

    pipedOutput: bool = not sys.stdout.isatty()
    fullOutput: bool = args.all or pipedOutput or args.filter or args.location or args.type or args.after or args.before
    if fullOutput:
        printObjectList(series.transactions, args.csv)
        exit(0)

    printObjectList(series.transactions[:20], args.csv)
    print(f'\n...<only showing 20>', file=sys.stderr)
    exit(0)